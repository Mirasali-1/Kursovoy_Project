from fastapi import FastAPI, Request, Form, Depends, HTTPException, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
from sqlalchemy.orm import Session
from sqlalchemy import desc
from datetime import datetime
import secrets
import os
import re
from pathlib import Path
from io import StringIO
import csv
from database import engine, get_db
from models import Base, User, Role, Product, Category, Zone, Operation, OperationItem, Supplier, Customer

Base.metadata.create_all(bind=engine)

app = FastAPI(
    title="WMS - Warehouse Management System",
    description="Система управления складскими операциями",
    version="2.0.0"
)

app.add_middleware(SessionMiddleware, secret_key=secrets.token_hex(32))
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")


def check_auth(request: Request):
    return request.session.get('logged_in', False)


def get_current_user(request: Request, db: Session):
    user_id = request.session.get('user_id')
    if not user_id:
        return None
    return db.query(User).filter(User.id == user_id).first()


def is_admin(user: User) -> bool:
    """Check if user has admin role"""
    if not user or not user.role:
        return False
    return user.role.name.lower() in ('admin', 'администратор', 'administrator', 'superadmin')


@app.get("/", response_class=HTMLResponse)
async def root(request: Request):
    if check_auth(request):
        return RedirectResponse(url="/dashboard", status_code=303)
    return RedirectResponse(url="/login", status_code=303)


@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    error = request.session.pop('error', None)
    return templates.TemplateResponse("login.html", {"request": request, "error": error})


@app.post("/login")
async def login_post(
        request: Request,
        email: str = Form(...),
        password: str = Form(...),
        db: Session = Depends(get_db)
):
    user = db.query(User).filter(User.email == email).first()

    if not user or user.password != password:
        request.session['error'] = "Неверный email или пароль"
        return RedirectResponse(url="/login", status_code=303)

    if not user.is_active:
        request.session['error'] = "Аккаунт деактивирован"
        return RedirectResponse(url="/login", status_code=303)

    request.session['logged_in'] = True
    request.session['user_id'] = user.id
    request.session['user_email'] = user.email

    return RedirectResponse(url="/dashboard", status_code=303)


@app.get("/logout")
async def logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/login", status_code=303)


@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard(
        request: Request,
        tab: str = "overview",
        operation_type: str = "all",
        status: str = "all",
        page: int = 1,
        db: Session = Depends(get_db)
):
    if not check_auth(request):
        return RedirectResponse(url="/login", status_code=303)

    current_user = get_current_user(request, db)
    if not current_user:
        request.session.clear()
        return RedirectResponse(url="/login", status_code=303)

    total_products = db.query(Product).filter(Product.is_active.is_not(False)).count()
    operations_today = db.query(Operation).filter(
        Operation.operation_date >= datetime.now().date()
    ).count()
    pending_operations = db.query(Operation).filter(
        Operation.status.in_(['pending', 'processing'])
    ).count()
    active_users = db.query(User).filter(User.is_active.is_not(False)).count()
    # NEW: total completed operations count
    total_completed = db.query(Operation).filter(Operation.status == 'completed').count()

    stats = [
        {'value': f"{total_products:,}", 'label': 'Всего товаров', 'icon': '📊', 'color': 'blue'},
        {'value': str(operations_today), 'label': 'Операций сегодня', 'icon': '✓', 'color': 'green'},
        {'value': str(pending_operations), 'label': 'Требуют внимания', 'icon': '⚠️', 'color': 'orange'},
        {'value': str(active_users), 'label': 'Активных сотрудников', 'icon': '👥', 'color': 'purple'}
    ]

    query = db.query(Operation)
    if operation_type != 'all':
        query = query.filter(Operation.operation_type == operation_type)
    if status != 'all':
        query = query.filter(Operation.status == status)

    per_page = 10
    total_ops = query.count()
    db_operations = query.order_by(desc(Operation.operation_date)).limit(per_page * page).all()
    has_more = total_ops > per_page * page

    # Count all operations (unfiltered) for the header counter
    all_ops_count = db.query(Operation).count()
    all_completed_count = db.query(Operation).filter(Operation.status == 'completed').count()

    operations = []
    for op in db_operations:
        status_map = {
            'completed': {'status': 'success', 'text': 'Завершено'},
            'processing': {'status': 'processing', 'text': 'В процессе'},
            'pending': {'status': 'pending', 'text': 'Ожидает'},
            'cancelled': {'status': 'pending', 'text': 'Отменено'}
        }
        status_info = status_map.get(op.status, {'status': 'pending', 'text': 'Неизвестно'})

        type_map = {
            'acceptance': 'Приемка товара',
            'shipment': 'Отгрузка заказа',
            'movement': 'Перемещение товара',
            'inventory': 'Инвентаризация'
        }
        title = type_map.get(op.operation_type, 'Операция')

        desc_parts = []
        if op.supplier_customer:
            desc_parts.append(op.supplier_customer)
        if op.notes:
            desc_parts.append(op.notes[:60] + ('...' if len(op.notes) > 60 else ''))
        desc_parts.append(op.operation_date.strftime('%d.%m.%Y, %H:%M'))

        # Check if current user can change the status of this operation
        can_manage = is_admin(current_user) or (op.user_id == current_user.id)

        operations.append({
            'db_id': op.id,
            'id': op.operation_number,
            'title': title,
            'desc': ' • '.join(desc_parts),
            'status': status_info['status'],
            'status_text': status_info['text'],
            'raw_status': op.status,
            'can_manage': can_manage,
        })

    user_data = {
        'name': current_user.full_name,
        'role': current_user.role.name if current_user.role else 'Не назначена',
        'initials': ''.join([word[0] for word in current_user.full_name.split()[:2]]),
        'email': current_user.email,
        'phone': current_user.phone or 'Не указан',
        'employee_id': current_user.employee_id,
        'avatar': current_user.avatar or None,
        'is_admin': is_admin(current_user),
    }

    success = request.session.pop('success', None)

    return templates.TemplateResponse("dashboard.html", {
        "request": request,
        "user": user_data,
        "current_tab": tab,
        "stats": stats,
        "operations": operations,
        "quick_actions": [],
        "has_more": has_more,
        "current_page": page,
        "operation_type": operation_type,
        "status": status,
        "success": success,
        "all_ops_count": all_ops_count,
        "all_completed_count": all_completed_count,
        "total_completed": total_completed,
    })


# ─── NEW: Change operation status ────────────────────────────────────────────

@app.post("/operations/{operation_id}/status")
async def change_operation_status(
        operation_id: int,
        request: Request,
        new_status: str = Form(...),
        db: Session = Depends(get_db)
):
    if not check_auth(request):
        return RedirectResponse(url="/login", status_code=303)

    current_user = get_current_user(request, db)
    if not current_user:
        request.session.clear()
        return RedirectResponse(url="/login", status_code=303)

    operation = db.query(Operation).filter(Operation.id == operation_id).first()
    if not operation:
        raise HTTPException(status_code=404, detail="Операция не найдена")

    # Only the operation owner or an admin can change status
    if not (is_admin(current_user) or operation.user_id == current_user.id):
        request.session['error'] = 'Нет прав для изменения статуса этой операции'
        return RedirectResponse(url="/dashboard?tab=operations", status_code=303)

    allowed_statuses = ['pending', 'processing', 'completed', 'cancelled']
    if new_status not in allowed_statuses:
        request.session['error'] = 'Недопустимый статус'
        return RedirectResponse(url="/dashboard?tab=operations", status_code=303)

    operation.status = new_status
    if new_status == 'completed' and not operation.completed_at:
        operation.completed_at = datetime.now()
    elif new_status != 'completed':
        operation.completed_at = None

    db.commit()

    status_names = {
        'pending': 'Ожидает',
        'processing': 'В процессе',
        'completed': 'Завершено',
        'cancelled': 'Отменено'
    }
    request.session['success'] = f'Статус операции {operation.operation_number} изменён на «{status_names[new_status]}»'
    return RedirectResponse(url="/dashboard?tab=operations", status_code=303)


# ─── NEW: Change status from details page ────────────────────────────────────

@app.post("/operations/details/{operation_id}/status")
async def change_operation_status_from_details(
        operation_id: int,
        request: Request,
        new_status: str = Form(...),
        db: Session = Depends(get_db)
):
    if not check_auth(request):
        return RedirectResponse(url="/login", status_code=303)

    current_user = get_current_user(request, db)
    if not current_user:
        request.session.clear()
        return RedirectResponse(url="/login", status_code=303)

    operation = db.query(Operation).filter(Operation.id == operation_id).first()
    if not operation:
        raise HTTPException(status_code=404, detail="Операция не найдена")

    if not (is_admin(current_user) or operation.user_id == current_user.id):
        request.session['error'] = 'Нет прав для изменения статуса этой операции'
        return RedirectResponse(url=f"/operations/details/{operation_id}", status_code=303)

    allowed_statuses = ['pending', 'processing', 'completed', 'cancelled']
    if new_status not in allowed_statuses:
        request.session['error'] = 'Недопустимый статус'
        return RedirectResponse(url=f"/operations/details/{operation_id}", status_code=303)

    operation.status = new_status
    if new_status == 'completed' and not operation.completed_at:
        operation.completed_at = datetime.now()
    elif new_status != 'completed':
        operation.completed_at = None

    db.commit()

    status_names = {
        'pending': 'Ожидает',
        'processing': 'В процессе',
        'completed': 'Завершено',
        'cancelled': 'Отменено'
    }
    request.session['success'] = f'Статус изменён на «{status_names[new_status]}»'
    return RedirectResponse(url=f"/operations/details/{operation_id}", status_code=303)


@app.get("/profile", response_class=HTMLResponse)
async def profile_page(request: Request, db: Session = Depends(get_db)):
    if not check_auth(request):
        return RedirectResponse(url="/login", status_code=303)

    current_user = get_current_user(request, db)
    if not current_user:
        request.session.clear()
        return RedirectResponse(url="/login", status_code=303)

    total_operations = db.query(Operation).filter(Operation.user_id == current_user.id).count()
    completed_operations = db.query(Operation).filter(
        Operation.user_id == current_user.id,
        Operation.status == 'completed'
    ).count()

    accuracy = round((completed_operations / total_operations * 100), 1) if total_operations > 0 else 0

    user_data = {
        'id': current_user.id,
        'name': current_user.full_name,
        'role': current_user.role.name if current_user.role else 'Не назначена',
        'initials': ''.join([word[0] for word in current_user.full_name.split()[:2]]),
        'email': current_user.email,
        'phone': current_user.phone or '',
        'employee_id': current_user.employee_id,
        'start_date': current_user.start_date.strftime('%d.%m.%Y') if current_user.start_date else '',
        'avatar': current_user.avatar or None,
        'total_operations': total_operations,
        'completed_operations': completed_operations,
        'accuracy': accuracy
    }

    success = request.session.pop('success', None)
    error = request.session.pop('error', None)

    return templates.TemplateResponse("profile.html", {
        "request": request,
        "user": user_data,
        "success": success,
        "error": error
    })


@app.post("/profile/update")
async def profile_update(
        request: Request,
        full_name: str = Form(...),
        email: str = Form(...),
        phone: str = Form(None),
        avatar: UploadFile = File(None),
        db: Session = Depends(get_db)
):
    if not check_auth(request):
        return RedirectResponse(url="/login", status_code=303)

    current_user = get_current_user(request, db)
    if not current_user:
        request.session.clear()
        return RedirectResponse(url="/login", status_code=303)

    current_user.full_name = full_name
    current_user.email = email
    current_user.phone = phone

    if avatar and avatar.filename:
        allowed_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.webp']
        file_ext = os.path.splitext(avatar.filename)[1].lower()

        if file_ext not in allowed_extensions:
            request.session['error'] = 'Неверный формат файла. Разрешены: JPG, PNG, GIF, WEBP'
            return RedirectResponse(url="/profile", status_code=303)

        filename = f"user_{current_user.id}_{secrets.token_hex(8)}{file_ext}"
        filepath = Path("static/avatars") / filename

        filepath.parent.mkdir(parents=True, exist_ok=True)

        if current_user.avatar:
            old_avatar_path = Path(current_user.avatar.lstrip('/'))
            if old_avatar_path.exists():
                old_avatar_path.unlink()

        with open(filepath, "wb") as f:
            content = await avatar.read()
            f.write(content)

        current_user.avatar = f"/static/avatars/{filename}"

    db.commit()

    request.session['success'] = 'Профиль успешно обновлён!'
    return RedirectResponse(url="/profile", status_code=303)


@app.post("/profile/delete-avatar")
async def delete_avatar(request: Request, db: Session = Depends(get_db)):
    if not check_auth(request):
        return RedirectResponse(url="/login", status_code=303)

    current_user = get_current_user(request, db)
    if not current_user:
        request.session.clear()
        return RedirectResponse(url="/login", status_code=303)

    if current_user.avatar:
        avatar_path = Path(current_user.avatar.lstrip('/'))
        if avatar_path.exists():
            avatar_path.unlink()

        current_user.avatar = None
        db.commit()

        request.session['success'] = 'Аватарка удалена'

    return RedirectResponse(url="/profile", status_code=303)


@app.get("/operations/details/{operation_id}", response_class=HTMLResponse)
async def operation_details(
        operation_id: int,
        request: Request,
        db: Session = Depends(get_db)
):
    if not check_auth(request):
        return RedirectResponse(url="/login", status_code=303)

    current_user = get_current_user(request, db)
    if not current_user:
        request.session.clear()
        return RedirectResponse(url="/login", status_code=303)

    operation = db.query(Operation).filter(Operation.id == operation_id).first()
    if not operation:
        raise HTTPException(status_code=404, detail="Операция не найдена")

    items = db.query(OperationItem).filter(OperationItem.operation_id == operation_id).all()

    type_map = {
        'acceptance': 'Приемка товара',
        'shipment': 'Отгрузка заказа',
        'movement': 'Перемещение товара',
        'inventory': 'Инвентаризация'
    }

    status_map = {
        'completed': {'status': 'success', 'text': 'Завершено'},
        'processing': {'status': 'processing', 'text': 'В процессе'},
        'pending': {'status': 'pending', 'text': 'Ожидает'},
        'cancelled': {'status': 'pending', 'text': 'Отменено'}
    }

    can_manage = is_admin(current_user) or (operation.user_id == current_user.id)

    operation_data = {
        'id': operation.id,
        'number': operation.operation_number,
        'type': type_map.get(operation.operation_type, 'Операция'),
        'status': status_map.get(operation.status, {'status': 'pending', 'text': 'Неизвестно'}),
        'raw_status': operation.status,
        'supplier_customer': operation.supplier_customer or '—',
        'total_amount': float(operation.total_amount) if operation.total_amount else 0,
        'notes': operation.notes or '—',
        'date': operation.operation_date.strftime('%d.%m.%Y %H:%M') if operation.operation_date else '—',
        'completed_at': operation.completed_at.strftime('%d.%m.%Y %H:%M') if operation.completed_at else '—',
        'user_name': operation.user.full_name if operation.user else 'Неизвестен',
        'can_manage': can_manage,
    }

    items_data = []
    for item in items:
        product = item.product
        items_data.append({
            'product_name': product.name if product else 'Товар удален',
            'quantity': item.quantity,
            'price': float(item.price) if item.price else 0,
            'total': item.quantity * float(item.price) if item.price else 0
        })

    user_data = {
        'name': current_user.full_name,
        'role': current_user.role.name if current_user.role else 'Не назначена',
        'initials': ''.join([word[0] for word in current_user.full_name.split()[:2]]),
        'avatar': current_user.avatar or None,
        'is_admin': is_admin(current_user),
    }

    success = request.session.pop('success', None)
    error = request.session.pop('error', None)

    return templates.TemplateResponse("operation_details.html", {
        "request": request,
        "user": user_data,
        "operation": operation_data,
        "items": items_data,
        "success": success,
        "error": error,
    })


@app.get("/operations/export")
async def export_operations(
        request: Request,
        operation_type: str = None,
        status: str = None,
        db: Session = Depends(get_db)
):
    if not check_auth(request):
        return RedirectResponse(url="/login", status_code=303)

    query = db.query(Operation)

    if operation_type and operation_type != 'all':
        query = query.filter(Operation.operation_type == operation_type)

    if status and status != 'all':
        query = query.filter(Operation.status == status)

    operations = query.order_by(desc(Operation.operation_date)).all()

    output = StringIO()
    writer = csv.writer(output)

    writer.writerow(['Номер', 'Тип', 'Статус', 'Поставщик/Заказчик', 'Сумма', 'Дата', 'Ответственный'])

    type_map = {
        'acceptance': 'Приемка',
        'shipment': 'Отгрузка',
        'movement': 'Перемещение',
        'inventory': 'Инвентаризация'
    }

    status_map = {
        'completed': 'Завершено',
        'processing': 'В процессе',
        'pending': 'Ожидает',
        'cancelled': 'Отменено'
    }

    for op in operations:
        writer.writerow([
            op.operation_number,
            type_map.get(op.operation_type, op.operation_type),
            status_map.get(op.status, op.status),
            op.supplier_customer or '',
            float(op.total_amount) if op.total_amount else 0,
            op.operation_date.strftime('%d.%m.%Y %H:%M') if op.operation_date else '',
            op.user.full_name if op.user else ''
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue().encode('utf-8-sig')]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=operations.csv"}
    )


@app.post("/api/users")
async def create_user(
        email: str = Form(...),
        password: str = Form(...),
        full_name: str = Form(...),
        phone: str = Form(None),
        role_id: int = Form(None),
        employee_id: str = Form(None),
        db: Session = Depends(get_db)
):
    existing = db.query(User).filter(User.email == email).first()
    if existing:
        raise HTTPException(status_code=400, detail="Пользователь с таким email уже существует")

    user = User(
        email=email,
        password=password,
        full_name=full_name,
        phone=phone,
        role_id=role_id,
        employee_id=employee_id
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return {"message": "Пользователь создан", "user_id": user.id}


@app.get("/api/products")
async def get_products(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    products = db.query(Product).filter(Product.is_active == True).offset(skip).limit(limit).all()
    return {
        "products": [{"id": p.id, "name": p.name, "quantity": p.quantity, "price": float(p.price)} for p in products]}


@app.get("/api/products/{product_id}")
async def get_product(product_id: int, db: Session = Depends(get_db)):
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Товар не найден")
    return {
        "id": product.id,
        "name": product.name,
        "article": product.article,
        "quantity": product.quantity,
        "price": float(product.price),
        "unit": product.unit
    }


@app.get("/api/operations")
async def get_operations(skip: int = 0, limit: int = 50, db: Session = Depends(get_db)):
    operations = db.query(Operation).order_by(desc(Operation.operation_date)).offset(skip).limit(limit).all()
    return {"operations": [
        {
            "id": op.id,
            "number": op.operation_number,
            "type": op.operation_type,
            "status": op.status,
            "date": op.operation_date.isoformat()
        } for op in operations
    ]}


@app.get("/api/operations/{operation_id}")
async def get_operation(operation_id: int, db: Session = Depends(get_db)):
    operation = db.query(Operation).filter(Operation.id == operation_id).first()
    if not operation:
        raise HTTPException(status_code=404, detail="Операция не найдена")
    return {
        "id": operation.id,
        "number": operation.operation_number,
        "type": operation.operation_type,
        "status": operation.status,
        "supplier_customer": operation.supplier_customer,
        "total_amount": float(operation.total_amount),
        "notes": operation.notes,
        "date": operation.operation_date.isoformat()
    }


@app.get("/api/categories")
async def get_categories(db: Session = Depends(get_db)):
    categories = db.query(Category).all()
    return {"categories": [{"id": c.id, "name": c.name, "description": c.description} for c in categories]}


@app.get("/api/zones")
async def get_zones(db: Session = Depends(get_db)):
    zones = db.query(Zone).filter(Zone.is_active == True).all()
    return {"zones": [{"id": z.id, "code": z.code, "name": z.name, "capacity": z.capacity} for z in zones]}


@app.get("/api/roles")
async def get_roles(db: Session = Depends(get_db)):
    roles = db.query(Role).all()
    return {"roles": [{"id": r.id, "name": r.name, "description": r.description} for r in roles]}


@app.get("/acceptance", response_class=HTMLResponse)
async def acceptance_page(request: Request, db: Session = Depends(get_db)):
    if not check_auth(request):
        return RedirectResponse(url="/login", status_code=303)

    current_user = get_current_user(request, db)
    if not current_user:
        request.session.clear()
        return RedirectResponse(url="/login", status_code=303)

    last_operation = db.query(Operation).filter(
        Operation.operation_type == 'acceptance'
    ).order_by(desc(Operation.id)).first()

    if last_operation:
        number = int(last_operation.operation_number.split('-')[1]) + 1
    else:
        number = 1

    next_number = f"ПР-{number:04d}"
    today = datetime.now().strftime('%Y-%m-%d')

    suppliers = db.query(Supplier).filter(Supplier.is_active.is_not(False)).all()
    products = db.query(Product).filter(Product.is_active.is_not(False)).all()
    zones = db.query(Zone).filter(Zone.is_active.is_not(False)).all()

    user_data = {
        'name': current_user.full_name,
        'role': current_user.role.name if current_user.role else 'Не назначена',
        'initials': ''.join([word[0] for word in current_user.full_name.split()[:2]]),
        'avatar': current_user.avatar or None
    }

    success = request.session.pop('success', None)
    error = request.session.pop('error', None)

    return templates.TemplateResponse("acceptance.html", {
        "request": request,
        "user": user_data,
        "next_number": next_number,
        "today": today,
        "suppliers": suppliers,
        "products": products,
        "zones": zones,
        "success": success,
        "error": error
    })


@app.post("/acceptance/create")
async def acceptance_create(
        request: Request,
        db: Session = Depends(get_db)
):
    if not check_auth(request):
        return RedirectResponse(url="/login", status_code=303)

    current_user = get_current_user(request, db)
    if not current_user:
        request.session.clear()
        return RedirectResponse(url="/login", status_code=303)

    form_data = await request.form()

    operation_number = form_data.get('operation_number')
    operation_date = form_data.get('operation_date')
    supplier = form_data.get('supplier')
    notes = form_data.get('notes')

    total_amount = 0
    products_data = {}

    for key in form_data.keys():
        if key.startswith('products['):
            m = re.match(r'products\[(\d+)\]\[(\w+)\]', key)
            if m:
                index, field = m.group(1), m.group(2)
                if index not in products_data:
                    products_data[index] = {}
                products_data[index][field] = form_data.get(key)

    for product_info in products_data.values():
        quantity = int(product_info.get('quantity', 0))
        price = float(product_info.get('price', 0))
        total_amount += quantity * price

    operation = Operation(
        operation_type='acceptance',
        operation_number=operation_number,
        status='completed',
        user_id=current_user.id,
        supplier_customer=supplier,
        total_amount=total_amount,
        notes=notes,
        operation_date=datetime.strptime(operation_date, '%Y-%m-%d'),
        completed_at=datetime.now()
    )

    db.add(operation)
    db.commit()
    db.refresh(operation)

    for product_info in products_data.values():
        product_id = int(product_info.get('product_id'))
        quantity = int(product_info.get('quantity'))
        price = float(product_info.get('price'))
        zone_id = int(product_info.get('zone_id'))

        operation_item = OperationItem(
            operation_id=operation.id,
            product_id=product_id,
            quantity=quantity,
            price=price,
            to_zone_id=zone_id
        )
        db.add(operation_item)

        product = db.query(Product).filter(Product.id == product_id).first()
        if product:
            product.quantity += quantity
            product.zone_id = zone_id

    db.commit()

    request.session['success'] = f'Приёмка {operation_number} успешно оформлена!'
    return RedirectResponse(url="/dashboard?tab=operations", status_code=303)


@app.get("/shipment", response_class=HTMLResponse)
async def shipment_page(request: Request, db: Session = Depends(get_db)):
    if not check_auth(request):
        return RedirectResponse(url="/login", status_code=303)

    current_user = get_current_user(request, db)
    if not current_user:
        request.session.clear()
        return RedirectResponse(url="/login", status_code=303)

    last_operation = db.query(Operation).filter(
        Operation.operation_type == 'shipment'
    ).order_by(desc(Operation.id)).first()

    if last_operation:
        number = int(last_operation.operation_number.split('-')[1]) + 1
    else:
        number = 1

    next_number = f"ОТ-{number:04d}"
    today = datetime.now().strftime('%Y-%m-%d')

    customers = db.query(Customer).filter(Customer.is_active.is_not(False)).all()
    products = db.query(Product).filter(
        Product.is_active.is_not(False),
        Product.quantity > 0
    ).all()

    user_data = {
        'name': current_user.full_name,
        'role': current_user.role.name if current_user.role else 'Не назначена',
        'initials': ''.join([word[0] for word in current_user.full_name.split()[:2]]),
        'avatar': current_user.avatar or None
    }

    success = request.session.pop('success', None)
    error = request.session.pop('error', None)

    return templates.TemplateResponse("shipment.html", {
        "request": request,
        "user": user_data,
        "next_number": next_number,
        "today": today,
        "customers": customers,
        "products": products,
        "success": success,
        "error": error
    })


@app.post("/shipment/create")
async def shipment_create(
        request: Request,
        db: Session = Depends(get_db)
):
    if not check_auth(request):
        return RedirectResponse(url="/login", status_code=303)

    current_user = get_current_user(request, db)
    if not current_user:
        request.session.clear()
        return RedirectResponse(url="/login", status_code=303)

    form_data = await request.form()

    operation_number = form_data.get('operation_number')
    operation_date = form_data.get('operation_date')
    customer = form_data.get('customer')
    delivery_address = form_data.get('delivery_address')
    notes = form_data.get('notes', '')

    if delivery_address:
        notes = f"Адрес доставки: {delivery_address}. {notes}"

    total_amount = 0
    products_data = {}

    for key in form_data.keys():
        if key.startswith('products['):
            m = re.match(r'products\[(\d+)\]\[(\w+)\]', key)
            if m:
                index, field = m.group(1), m.group(2)
                if index not in products_data:
                    products_data[index] = {}
                products_data[index][field] = form_data.get(key)

    for product_info in products_data.values():
        quantity = int(product_info.get('quantity', 0))
        price = float(product_info.get('price', 0))
        total_amount += quantity * price

    operation = Operation(
        operation_type='shipment',
        operation_number=operation_number,
        status='completed',
        user_id=current_user.id,
        supplier_customer=customer,
        total_amount=total_amount,
        notes=notes,
        operation_date=datetime.strptime(operation_date, '%Y-%m-%d'),
        completed_at=datetime.now()
    )

    db.add(operation)
    db.commit()
    db.refresh(operation)

    for product_info in products_data.values():
        product_id = int(product_info.get('product_id'))
        quantity = int(product_info.get('quantity'))
        price = float(product_info.get('price'))

        operation_item = OperationItem(
            operation_id=operation.id,
            product_id=product_id,
            quantity=quantity,
            price=price,
            from_zone_id=None
        )
        db.add(operation_item)

        product = db.query(Product).filter(Product.id == product_id).first()
        if product:
            if product.quantity >= quantity:
                product.quantity -= quantity
            else:
                request.session['error'] = f'Недостаточно товара {product.name} на складе!'
                db.rollback()
                return RedirectResponse(url="/shipment", status_code=303)

    db.commit()

    request.session['success'] = f'Отгрузка {operation_number} успешно оформлена!'
    return RedirectResponse(url="/dashboard?tab=operations", status_code=303)


@app.get("/movement", response_class=HTMLResponse)
async def movement_page(request: Request, db: Session = Depends(get_db)):
    if not check_auth(request):
        return RedirectResponse(url="/login", status_code=303)

    current_user = get_current_user(request, db)
    if not current_user:
        request.session.clear()
        return RedirectResponse(url="/login", status_code=303)

    last_operation = db.query(Operation).filter(
        Operation.operation_type == 'movement'
    ).order_by(desc(Operation.id)).first()

    if last_operation:
        try:
            number = int(last_operation.operation_number.split('-')[1]) + 1
        except (IndexError, ValueError):
            number = 1
    else:
        number = 1

    next_number = f"ПМ-{number:04d}"
    today = datetime.now().strftime('%Y-%m-%d')

    products = db.query(Product).filter(
        Product.is_active.is_not(False),
        Product.quantity > 0
    ).all()
    zones = db.query(Zone).filter(Zone.is_active.is_not(False)).all()

    user_data = {
        'name': current_user.full_name,
        'role': current_user.role.name if current_user.role else 'Не назначена',
        'initials': ''.join([word[0] for word in current_user.full_name.split()[:2]]),
        'avatar': current_user.avatar or None
    }

    success = request.session.pop('success', None)
    error = request.session.pop('error', None)

    return templates.TemplateResponse("movement.html", {
        "request": request,
        "user": user_data,
        "next_number": next_number,
        "today": today,
        "products": products,
        "zones": zones,
        "success": success,
        "error": error
    })


@app.post("/movement/create")
async def movement_create(request: Request, db: Session = Depends(get_db)):
    if not check_auth(request):
        return RedirectResponse(url="/login", status_code=303)

    current_user = get_current_user(request, db)
    if not current_user:
        request.session.clear()
        return RedirectResponse(url="/login", status_code=303)

    form_data = await request.form()

    operation_number = form_data.get('operation_number')
    operation_date = form_data.get('operation_date')
    notes = form_data.get('notes', '')

    products_data = {}
    for key in form_data.keys():
        if key.startswith('products['):
            match = re.match(r'products\[(\d+)\]\[(\w+)\]', key)
            if match:
                index = match.group(1)
                field = match.group(2)
                if index not in products_data:
                    products_data[index] = {}
                products_data[index][field] = form_data.get(key)

    if not products_data:
        request.session['error'] = 'Необходимо добавить хотя бы один товар'
        return RedirectResponse(url="/movement", status_code=303)

    for product_info in products_data.values():
        product_id = int(product_info.get('product_id', 0))
        quantity = int(product_info.get('quantity', 0))
        from_zone_id = int(product_info.get('from_zone_id', 0))
        to_zone_id = int(product_info.get('to_zone_id', 0))

        if from_zone_id == to_zone_id:
            request.session['error'] = 'Зона "откуда" и "куда" не могут совпадать!'
            return RedirectResponse(url="/movement", status_code=303)

        product = db.query(Product).filter(Product.id == product_id).first()
        if product and product.quantity < quantity:
            request.session['error'] = f'Недостаточно товара "{product.name}"! Доступно: {product.quantity}'
            return RedirectResponse(url="/movement", status_code=303)

    operation = Operation(
        operation_type='movement',
        operation_number=operation_number,
        status='completed',
        user_id=current_user.id,
        supplier_customer=None,
        total_amount=0,
        notes=notes,
        operation_date=datetime.strptime(operation_date, '%Y-%m-%d'),
        completed_at=datetime.now()
    )
    db.add(operation)
    db.commit()
    db.refresh(operation)

    for product_info in products_data.values():
        product_id = int(product_info.get('product_id', 0))
        quantity = int(product_info.get('quantity', 0))
        from_zone_id = int(product_info.get('from_zone_id', 0))
        to_zone_id = int(product_info.get('to_zone_id', 0))

        operation_item = OperationItem(
            operation_id=operation.id,
            product_id=product_id,
            quantity=quantity,
            price=0,
            from_zone_id=from_zone_id,
            to_zone_id=to_zone_id
        )
        db.add(operation_item)

        product = db.query(Product).filter(Product.id == product_id).first()
        if product:
            product.zone_id = to_zone_id

    db.commit()

    request.session['success'] = f'Перемещение {operation_number} успешно оформлено!'
    return RedirectResponse(url="/dashboard?tab=operations", status_code=303)


@app.get("/inventory", response_class=HTMLResponse)
async def inventory_page(request: Request, db: Session = Depends(get_db)):
    if not check_auth(request):
        return RedirectResponse(url="/login", status_code=303)

    current_user = get_current_user(request, db)
    if not current_user:
        request.session.clear()
        return RedirectResponse(url="/login", status_code=303)

    last_operation = db.query(Operation).filter(
        Operation.operation_type == 'inventory'
    ).order_by(desc(Operation.id)).first()

    if last_operation:
        try:
            number = int(last_operation.operation_number.split('-')[1]) + 1
        except (IndexError, ValueError):
            number = 1
    else:
        number = 1

    next_number = f"ИН-{number:04d}"
    today = datetime.now().strftime('%Y-%m-%d')

    products = db.query(Product).filter(Product.is_active.is_not(False)).all()
    zones = db.query(Zone).filter(Zone.is_active.is_not(False)).all()

    user_data = {
        'name': current_user.full_name,
        'role': current_user.role.name if current_user.role else 'Не назначена',
        'initials': ''.join([word[0] for word in current_user.full_name.split()[:2]]),
        'avatar': current_user.avatar or None
    }

    success = request.session.pop('success', None)
    error = request.session.pop('error', None)

    return templates.TemplateResponse("inventory.html", {
        "request": request,
        "user": user_data,
        "next_number": next_number,
        "today": today,
        "products": products,
        "zones": zones,
        "success": success,
        "error": error
    })


@app.post("/inventory/create")
async def inventory_create(request: Request, db: Session = Depends(get_db)):
    if not check_auth(request):
        return RedirectResponse(url="/login", status_code=303)

    current_user = get_current_user(request, db)
    if not current_user:
        request.session.clear()
        return RedirectResponse(url="/login", status_code=303)

    form_data = await request.form()

    operation_number = form_data.get('operation_number')
    operation_date = form_data.get('operation_date')
    zone_filter = form_data.get('zone_filter', '')
    notes = form_data.get('notes', '')

    products_data = {}
    for key in form_data.keys():
        if key.startswith('products['):
            match = re.match(r'products\[(\d+)\]\[(\w+)\]', key)
            if match:
                index = match.group(1)
                field = match.group(2)
                if index not in products_data:
                    products_data[index] = {}
                products_data[index][field] = form_data.get(key)

    if not products_data:
        request.session['error'] = 'Нет данных для инвентаризации'
        return RedirectResponse(url="/inventory", status_code=303)

    discrepancies = []
    for product_info in products_data.values():
        product_id = int(product_info.get('product_id', 0))
        actual_qty = int(product_info.get('actual_quantity', 0))
        product = db.query(Product).filter(Product.id == product_id).first()
        if product and product.quantity != actual_qty:
            discrepancies.append(f"{product.name}: было {product.quantity}, стало {actual_qty}")

    if zone_filter:
        notes_full = f"Зона: {zone_filter}. {notes}" if notes else f"Зона: {zone_filter}."
    else:
        notes_full = notes

    if discrepancies:
        notes_full += " Расхождения: " + "; ".join(discrepancies)

    operation = Operation(
        operation_type='inventory',
        operation_number=operation_number,
        status='completed',
        user_id=current_user.id,
        supplier_customer=None,
        total_amount=0,
        notes=notes_full,
        operation_date=datetime.strptime(operation_date, '%Y-%m-%d'),
        completed_at=datetime.now()
    )
    db.add(operation)
    db.commit()
    db.refresh(operation)

    for product_info in products_data.values():
        product_id = int(product_info.get('product_id', 0))
        actual_qty = int(product_info.get('actual_quantity', 0))

        product = db.query(Product).filter(Product.id == product_id).first()
        if product:
            diff = actual_qty - product.quantity
            operation_item = OperationItem(
                operation_id=operation.id,
                product_id=product_id,
                quantity=abs(diff) if diff != 0 else actual_qty,
                price=0,
                from_zone_id=None,
                to_zone_id=None
            )
            db.add(operation_item)
            product.quantity = actual_qty

    db.commit()

    request.session['success'] = f'Инвентаризация {operation_number} успешно завершена! Расхождений: {len(discrepancies)}'
    return RedirectResponse(url="/dashboard?tab=operations", status_code=303)


@app.get("/reports", response_class=HTMLResponse)
async def reports_page(request: Request, db: Session = Depends(get_db)):
    if not check_auth(request):
        return RedirectResponse(url="/login", status_code=303)
    current_user = get_current_user(request, db)
    if not current_user:
        request.session.clear()
        return RedirectResponse(url="/login", status_code=303)

    user_data = {
        'name': current_user.full_name,
        'role': current_user.role.name if current_user.role else 'Не назначена',
        'initials': ''.join([word[0] for word in current_user.full_name.split()[:2]]),
        'avatar': current_user.avatar or None
    }
    return templates.TemplateResponse("reports.html", {"request": request, "user": user_data})


@app.get("/reports/export")
async def reports_export(
        request: Request,
        fmt: str = "csv",
        operation_type: str = "all",
        status: str = "all",
        date_from: str = None,
        date_to: str = None,
        db: Session = Depends(get_db)
):
    if not check_auth(request):
        return RedirectResponse(url="/login", status_code=303)

    query = db.query(Operation)
    if operation_type != 'all':
        query = query.filter(Operation.operation_type == operation_type)
    if status != 'all':
        query = query.filter(Operation.status == status)
    if date_from:
        query = query.filter(Operation.operation_date >= datetime.strptime(date_from, '%Y-%m-%d'))
    if date_to:
        query = query.filter(Operation.operation_date <= datetime.strptime(date_to, '%Y-%m-%d').replace(hour=23, minute=59))

    operations = query.order_by(desc(Operation.operation_date)).all()

    type_map = {'acceptance': 'Приемка', 'shipment': 'Отгрузка', 'movement': 'Перемещение', 'inventory': 'Инвентаризация'}
    status_map = {'completed': 'Завершено', 'processing': 'В процессе', 'pending': 'Ожидает', 'cancelled': 'Отменено'}

    rows = []
    for op in operations:
        items = db.query(OperationItem).filter(OperationItem.operation_id == op.id).all()
        items_desc = '; '.join([
            f"{item.product.name if item.product else '—'} x{item.quantity}"
            for item in items
        ])
        rows.append({
            'number': op.operation_number,
            'type': type_map.get(op.operation_type, op.operation_type),
            'status': status_map.get(op.status, op.status),
            'counterpart': op.supplier_customer or '—',
            'total': f"{float(op.total_amount):.2f}" if op.total_amount else '0.00',
            'items': items_desc,
            'notes': op.notes or '—',
            'date': op.operation_date.strftime('%d.%m.%Y %H:%M') if op.operation_date else '—',
            'user': op.user.full_name if op.user else '—',
        })

    if fmt == 'csv':
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(['Номер', 'Тип', 'Статус', 'Контрагент', 'Сумма (₽)', 'Товары', 'Примечания', 'Дата', 'Ответственный'])
        for r in rows:
            writer.writerow([r['number'], r['type'], r['status'], r['counterpart'], r['total'], r['items'], r['notes'], r['date'], r['user']])
        output.seek(0)
        return StreamingResponse(iter([output.getvalue().encode('utf-8-sig')]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=report.csv"})

    elif fmt == 'txt':
        lines = ["=" * 70, "ОТЧЁТ ПО СКЛАДСКИМ ОПЕРАЦИЯМ WMS", f"Сформирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}", "=" * 70, ""]
        for r in rows:
            lines += [
                f"Номер:        {r['number']}",
                f"Тип:          {r['type']}",
                f"Статус:       {r['status']}",
                f"Контрагент:   {r['counterpart']}",
                f"Сумма:        {r['total']} ₽",
                f"Товары:       {r['items']}",
                f"Примечания:   {r['notes']}",
                f"Дата:         {r['date']}",
                f"Ответственный:{r['user']}",
                "-" * 70, ""
            ]
        lines.append(f"Итого операций: {len(rows)}")
        content = "\n".join(lines)
        return StreamingResponse(iter([content.encode('utf-8')]),
            media_type="text/plain",
            headers={"Content-Disposition": "attachment; filename=report.txt"})

    elif fmt == 'xlsx':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from io import BytesIO

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Операции"

            # Header style
            header_fill = PatternFill("solid", fgColor="667EEA")
            header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
            border_side = Side(style='thin', color='DDDDDD')
            cell_border = Border(
                left=border_side, right=border_side,
                top=border_side, bottom=border_side
            )

            headers = ['Номер', 'Тип', 'Статус', 'Контрагент', 'Сумма (₽)', 'Товары', 'Примечания', 'Дата', 'Ответственный']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = cell_border

            ws.row_dimensions[1].height = 24

            alt_fill = PatternFill("solid", fgColor="F5F7FA")
            normal_font = Font(name="Calibri", size=10)

            for row_idx, r in enumerate(rows, 2):
                # Parse total safely
                try:
                    total_val = float(r['total'].replace(',', '.'))
                except (ValueError, AttributeError):
                    total_val = 0.0

                values = [
                    r['number'],
                    r['type'],
                    r['status'],
                    r['counterpart'],
                    total_val,
                    r['items'],
                    r['notes'],
                    r['date'],
                    r['user']
                ]
                fill = alt_fill if row_idx % 2 == 0 else None
                for col, val in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col, value=val)
                    cell.border = cell_border
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    cell.font = normal_font
                    if fill:
                        cell.fill = fill

            # Column widths
            col_widths = [14, 16, 14, 24, 12, 40, 30, 18, 20]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

            # Freeze header row
            ws.freeze_panes = 'A2'

            # Summary sheet
            ws2 = wb.create_sheet("Сводка")
            ws2['A1'] = 'Сводка по типам операций'
            ws2['A1'].font = Font(bold=True, size=14, name="Calibri")
            ws2['A2'] = f'Сформирован: {datetime.now().strftime("%d.%m.%Y %H:%M")}'
            ws2['A2'].font = Font(italic=True, color="888888", name="Calibri")

            type_counts = {}
            type_totals = {}
            for r in rows:
                t = r['type']
                type_counts[t] = type_counts.get(t, 0) + 1
                try:
                    type_totals[t] = type_totals.get(t, 0) + float(r['total'].replace(',', '.'))
                except (ValueError, AttributeError):
                    pass

            ws2.append([])
            header_row = ws2.max_row + 1
            ws2.append(['Тип операции', 'Количество', 'Сумма (₽)'])
            for cell in ws2[header_row]:
                cell.font = Font(bold=True, name="Calibri")
                cell.fill = PatternFill("solid", fgColor="E8F0FE")
                cell.border = cell_border

            for t, cnt in type_counts.items():
                ws2.append([t, cnt, round(type_totals.get(t, 0), 2)])

            total_row = ws2.max_row + 1
            n = len(type_counts)
            ws2.append(['ИТОГО', f'=SUM(B{header_row+1}:B{header_row+n})', f'=SUM(C{header_row+1}:C{header_row+n})'])
            for cell in ws2[total_row]:
                cell.font = Font(bold=True, name="Calibri")
                cell.fill = PatternFill("solid", fgColor="667EEA")
                cell.font = Font(bold=True, color="FFFFFF", name="Calibri")

            ws2.column_dimensions['A'].width = 22
            ws2.column_dimensions['B'].width = 14
            ws2.column_dimensions['C'].width = 16

            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            content = buf.read()

            return StreamingResponse(
                iter([content]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=report.xlsx"}
            )
        except ImportError:
            return StreamingResponse(
                iter([b"openpyxl not installed. Run: pip install openpyxl"]),
                media_type="text/plain",
                headers={"Content-Disposition": "attachment; filename=error.txt"}
            )

    elif fmt == 'pdf':
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import cm
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from io import BytesIO

            buf = BytesIO()
            doc = SimpleDocTemplate(
                buf,
                pagesize=landscape(A4),
                leftMargin=1.5*cm, rightMargin=1.5*cm,
                topMargin=2*cm, bottomMargin=2*cm
            )

            styles = getSampleStyleSheet()
            story = []

            # Use built-in Helvetica (no CJK/Cyrillic issues with latin subset)
            # For Cyrillic we wrap text in Paragraph with proper encoding
            title_style = ParagraphStyle(
                'WMSTitle',
                parent=styles['Normal'],
                fontSize=16,
                fontName='Helvetica-Bold',
                spaceAfter=6,
                textColor=colors.HexColor('#333333')
            )
            sub_style = ParagraphStyle(
                'WMSSub',
                parent=styles['Normal'],
                fontSize=10,
                fontName='Helvetica',
                textColor=colors.grey,
                spaceAfter=20
            )
            cell_style = ParagraphStyle(
                'WMSCell',
                parent=styles['Normal'],
                fontSize=8,
                fontName='Helvetica',
                leading=10,
            )

            story.append(Paragraph("WMS - Otchet po skladskim operatsiyam", title_style))
            story.append(Paragraph(
                f"Sformirovan: {datetime.now().strftime('%d.%m.%Y %H:%M')} | Operatsiy: {len(rows)}",
                sub_style
            ))
            story.append(Spacer(1, 0.3*cm))

            # Build table with Paragraphs for wrapping
            table_data = [[
                Paragraph('Nomer', cell_style),
                Paragraph('Tip', cell_style),
                Paragraph('Status', cell_style),
                Paragraph('Kontragent', cell_style),
                Paragraph('Summa', cell_style),
                Paragraph('Data', cell_style),
                Paragraph('Otvetstvenny', cell_style),
            ]]

            for r in rows:
                table_data.append([
                    Paragraph(str(r['number']), cell_style),
                    Paragraph(str(r['type']), cell_style),
                    Paragraph(str(r['status']), cell_style),
                    Paragraph(str(r['counterpart'])[:30], cell_style),
                    Paragraph(str(r['total']) + ' rub', cell_style),
                    Paragraph(str(r['date']), cell_style),
                    Paragraph(str(r['user'])[:25], cell_style),
                ])

            col_widths = [3.5*cm, 3.5*cm, 3*cm, 5*cm, 3*cm, 4*cm, 4.5*cm]
            t = Table(table_data, colWidths=col_widths, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667EEA')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F7FA')]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(t)

            doc.build(story)
            buf.seek(0)
            content = buf.read()

            return StreamingResponse(
                iter([content]),
                media_type="application/pdf",
                headers={"Content-Disposition": "attachment; filename=report.pdf"}
            )
        except ImportError:
            return StreamingResponse(
                iter([b"reportlab not installed. Run: pip install reportlab"]),
                media_type="text/plain",
                headers={"Content-Disposition": "attachment; filename=error.txt"}
            )

    elif fmt == 'json':
        import json
        content = json.dumps(
            {"generated_at": datetime.now().isoformat(), "total": len(rows), "operations": rows},
            ensure_ascii=False, indent=2
        )
        return StreamingResponse(iter([content.encode('utf-8')]),
            media_type="application/json",
            headers={"Content-Disposition": "attachment; filename=report.json"})

    elif fmt == 'html_report':
        html_rows = ""
        for i, r in enumerate(rows):
            bg = '#f5f7fa' if i % 2 == 0 else 'white'
            html_rows += (
                f"<tr style='background:{bg}'>"
                f"<td>{r['number']}</td><td>{r['type']}</td><td>{r['status']}</td>"
                f"<td>{r['counterpart']}</td><td>{r['total']} ₽</td>"
                f"<td>{r['items']}</td><td>{r['date']}</td><td>{r['user']}</td>"
                f"</tr>"
            )
        html = f"""<!DOCTYPE html><html lang="ru"><head><meta charset="UTF-8">
        <title>Отчёт WMS</title>
        <style>body{{font-family:Arial,sans-serif;margin:40px;color:#333}}h1{{color:#667eea}}table{{width:100%;border-collapse:collapse;margin-top:20px}}th{{background:#667eea;color:white;padding:10px;text-align:left}}td{{padding:8px 10px;border-bottom:1px solid #eee}}p{{color:#888}}</style></head>
        <body><h1>Отчёт по складским операциям WMS</h1>
        <p>Сформирован: {datetime.now().strftime('%d.%m.%Y %H:%M')} | Операций: {len(rows)}</p>
        <table><thead><tr><th>Номер</th><th>Тип</th><th>Статус</th><th>Контрагент</th><th>Сумма</th><th>Товары</th><th>Дата</th><th>Ответственный</th></tr></thead>
        <tbody>{html_rows}</tbody></table></body></html>"""
        return StreamingResponse(iter([html.encode('utf-8')]),
            media_type="text/html",
            headers={"Content-Disposition": "attachment; filename=report.html"})

    return RedirectResponse(url="/reports", status_code=303)


if __name__ == "__main__":
    import uvicorn

    print("=" * 60)
    print("🚀 Запуск WMS с PostgreSQL")
    print("=" * 60)
    print(f"🌐 Приложение: http://localhost:8000")
    print(f"📚 API документация: http://localhost:8000/docs")
    print(f"🗄️  База данных: PostgreSQL (wms_db)")
    print("")
    print("👤 Тестовый вход:")
    print("   Email: admin@wms.com")
    print("   Пароль: admin123")
    print("=" * 60)
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)